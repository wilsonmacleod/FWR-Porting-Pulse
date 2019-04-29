import openpyxl
import datetime
from openpyxl.utils import get_column_letter
import matplotlib
import matplotlib.pyplot as plt
from matplotlib import pyplot as mp
import os

class Porting_puls():
    
    def __init__(self, month_total = 0, ytd_value = 0):

        self.wb = openpyxl.load_workbook('porting_pulse.xlsx')
        self.sheet1 = self.wb['data2019']
        self.date_time = datetime.datetime.today().strftime('%Y-%month_value-%d')
        self.month_total = month_total
        self.ytd_value = ytd_value

    def find_col(self):

        """find column based on month"""
        month = self.date_time[5:7] #find the month by indexing datetime of today
        sheet_dict = {'01': '2', '02': '4', '03': '6', '04': '8', '05': '10', '06': '12', '07': '14', '08': '16', '09': '18', '10': '20', '11': '22', '12': '24'}
        col = sheet_dict[month]
        return col

    
    def save(self):

        """save our workbook"""

        self.wb.save('porting_pulse.xlsx')

    def write_um(self, port_count):

        """write port_count into the correct column(month) and row(week) of our workbook, port count from get.py"""

        col = self.find_col()
        row_variable = 3
        search_rows = 0
        month_value = self.sheet1.cell(row = 1, column = int(col)).value
        print(f'{month_value}.')

        for weekz in range(3,8): #set nonetypes to zero for later addition
            if self.sheet1.cell(row=weekz, column=int(col)).value == None:
                self.sheet1.cell(row=weekz, column=int(col)).value = 0
        
        while search_rows < 1: #search rows in month column to find correct cell and then write into it
            if self.sheet1.cell(row = int(row_variable), column = int(col)).value != 0:
                row_variable += 1
                continue
            elif row_variable >= 8:
                print('All weeks for this month are full please double check the spreadsheet.')
                break
            else:
                self.sheet1.cell(row = int(row_variable), column = int(col)).value = port_count
                search_rows += 1
        self.save()

    def write_vip(self, vip_dict):
            
            """write pon count per vip_dict(passed from get.py)"""

            col = self.find_col()

            for vip_cell in range(11,21): #set nonetypes to zero 
                if self.sheet1.cell(row=vip_cell, column=int(col)).value == None:
                    self.sheet1.cell(row=vip_cell, column=int(col)).value = 0
            
            vip_cell = 11 #vip_cell to start script at first vip
            for vip,value in vip_dict.items():
                if value != None:
                    self.sheet1.cell(row=vip_cell, column=int(col)).value += value
                    vip_cell +=1

            self.save()


    def mtd(self):

        """add value of rows for the month and return and write total"""

        col = self.find_col()
        self.month_total = 0
        week1 = self.sheet1.cell(row = 3, column = int(col)).value 
        week2 = self.sheet1.cell(row = 4, column = int(col)).value
        week3 = self.sheet1.cell(row = 5, column = int(col)).value
        week4 = self.sheet1.cell(row = 6, column = int(col)).value 
        week5 = self.sheet1.cell(row = 7, column = int(col)).value
        self.month_total = int(week1) + int(week2) + int(week3) + int(week4) + int(week5)
        
        print(f'MTD: {self.month_total}')
        self.sheet1.cell(row = 8, column = int(col)).value = int(self.month_total)
        self.save()
        return self.month_total
        
    def ytd(self):

        """find and write/update ytd cell with current ytd value"""
    
        col = self.find_col()
        pre_col = int(col) - 2 #find previous ytd cell

        if self.sheet1.cell(row = 9, column = int(pre_col)).value == None:
            self.sheet1.cell(row = 9, column = int(pre_col)).value = 0

        current_ytd = self.sheet1.cell(row = 9, column = int(pre_col)).value

        self.ytd_value = current_ytd + self.month_total
        self.sheet1.cell(row = 9, column = int(col)).value = int(self.ytd_value)
        self.save()
        return self.ytd_value

    def report_plot(self):

        """plot ytd total by week"""

        col = self.find_col() #find current column/month
        col = int(col)
        month_value = self.sheet1.cell(row = 1, column = int(col)-2).value #month for file naming
        by_week_list = [] #list we will append using nested while
        while col >= 2: #keep going until Jan
            
            row_num = 3 #begin at week1
            
            for row in range(3,8):
                
                week=self.sheet1.cell(row = int(row_num), column = int(col)).value
                
                if week != None and week != 0:
                    
                    by_week_list.append(week)
                    row_num+=1
            
            col -= 2 #after scanning all weeks move back a month
        
        os.makedirs(f'PortReport{month_value}', exist_ok=True) #make directort for results

        total_sum = 0
        for num in by_week_list: #find sum and total
            total_sum += num

        avg_ports = total_sum/len(by_week_list)

        print(f'YTD total port-ins:{total_sum}\nYTD average ports per week: {round(avg_ports,2)}')

        plt.plot([i for i in by_week_list], 'b-') #generate graph line graph for pulse on incoming port ins
        plt.axis([1,int(len(by_week_list)), 0, 300])
        plt.title(f'Port-In Pulse, YTD:{total_sum}\nAverage ports per week: {round(avg_ports,2)}')
        plt.ylabel('Port Count')
        plt.xlabel('Weeks YTD')
        mp.savefig(f'PortReport{month_value}/{month_value}_YTD.pdf')

    def report_month(self):

        """report monthly total and % increase or decrease"""
        
        col = self.find_col()
        col = int(col)-2 #will be running this part for a monthly report-need to start on the past months column
        month_value = self.sheet1.cell(row = 1, column = int(col)).value #month name for data presentation
        
        mtd_value = self.sheet1.cell(row = 8, column = int(col)).value
        mtd_value_past = self.sheet1.cell(row = 8, column = int(col)-2).value
        mtd_trend=((mtd_value-mtd_value_past)/mtd_value)*100 #find mtd trend

        print(f'\n--\n{month_value} total: {mtd_value}\nTrend compared to last month:{round(mtd_trend,2)}%\n--\n')

    def report_vip_comp(self):
        
        """find trend of vip porting activites"""

        col = self.find_col()
        col = int(col)-2#will be running this part for a monthly report-need to start on the past months column
        month_value = self.sheet1.cell(row = 1, column = int(col)).value
        
        trend_tup = ()
        names_list = []
        row_num = 11 #begin at vip1
        for each in names_list:

            vip_val=self.sheet1.cell(row = int(row_num), column = int(col)).value
            vip_val_past=self.sheet1.cell(row = int(row_num), column = int(col)-2).value

            if vip_val != None and vip_val != 0:
                trend=((vip_val-vip_val_past)/vip_val)*100 #find trend
                trend = (str(f'Previous Month: {vip_val_past},{month_value}: {vip_val}, Trend: {round(trend,2)}%'))#make more presentable
                trend_tup += (each,trend)#pair VIP name with % increase or decrease
            row_num+=1
    
        for each_one in trend_tup:
            print(each_one)




